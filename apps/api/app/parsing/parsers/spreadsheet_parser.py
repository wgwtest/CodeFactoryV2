from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from app.parsing.models import ParsedDocument, ParsedSegment
from app.parsing.parsers.doc_converter import convert_office_document


def parse_spreadsheet_segments(file_path: str) -> ParsedDocument:
    path = Path(file_path)
    if path.suffix.lower() == ".xls":
        with TemporaryDirectory() as temp_dir:
            converted = Path(convert_office_document(file_path, temp_dir, "xlsx"))
            return _parse_xlsx_segments(converted)

    return _parse_xlsx_segments(path)


def _parse_xlsx_segments(path: Path) -> ParsedDocument:
    workbook = load_workbook(path, data_only=True)
    segments: list[ParsedSegment] = []
    for sheet in workbook.worksheets:
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            cells = [str(cell).strip() for cell in row if cell not in (None, "")]
            if not cells:
                continue
            heading = cells[0][:255]
            segments.append(
                ParsedSegment(
                    block_id=f"{sheet.title}-{row_index}",
                    heading=heading,
                    content=" | ".join(cells),
                    anchor={"sheet": sheet.title, "row": row_index, "section": heading},
                    block_type="table_row",
                )
            )
    return ParsedDocument(parser_name="spreadsheet_openpyxl", parser_version="v3", segments=segments)
